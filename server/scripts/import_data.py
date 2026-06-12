"""
Import historical data from 莫非凡讲规划-2026夏季高考数据库.xlsx into SQLite.

Data Source:
  C:/Users/wangm/Desktop/莫非凡讲规划-2026夏季高考数据库.xlsx

  Sheet "总表" — 33844 rows of 广东高考招生与录取数据

What this script imports:
  1. enrollment_plan  ← 2025 招生计划 (each row = one major in a group)
  2. admission_history ← 2025 专业组录取 (group-level, cols 24-26)
  3. admission_history ← 2024 专业录取 (aggregated to group-level from cols 35-37)

Usage:
  cd server
  python scripts/import_data.py
"""

import csv
import sqlite3
from collections import defaultdict
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────
EXCEL_PATH = Path(
    "C:/Users/wangm/Desktop/莫非凡讲规划-2026夏季高考数据库.xlsx"
)
DB_PATH = Path(__file__).resolve().parent.parent / "gaokao.db"
BACKUP_PATH = DB_PATH.with_suffix(".db.backup")

# ── Column indices (0-based) ──────────────────────────────────────────
# Basic plan info
C_YEAR = 0
C_PROVINCE = 1
C_BATCH = 2  # 批次: 本科批, 本科提前批, 专科批, etc.
C_SUBJECT = 4  # 科类: 物理 / 历史
C_SCHOOL_CODE = 5
C_SCHOOL_NAME = 6
C_GROUP_CODE = 7  # 院校专业组代码
C_MAJOR_CODE = 9
C_MAJOR_NAME = 11
C_MAJOR_NOTE = 12
C_SUBJECT_REQ = 15  # 选科要求
C_PLAN_COUNT = 16  # 计划人数 (单个专业)
C_GROUP_MAJORS = 19  # 组内专业列表
C_GROUP_PLAN = 20  # 专业组计划人数

# 2025 group-level admission (专业组录取)
C_G25_ADMIT = 23
C_G25_MIN_SCORE = 24
C_G25_MIN_RANK = 25

# 2025 major-level admission (专业录取)
C_M25_ADMIT = 26
C_M25_MIN_SCORE = 27
C_M25_MIN_RANK = 28

# 2024 major-level admission
C_M24_ADMIT = 34  # 录取人数2
C_M24_MIN_SCORE = 35  # 最低分2
C_M24_MIN_RANK = 36  # 最低位次2
C_M24_PLAN = 37  # 计划人数结果2

# School info
C_CITY = 39  # 城市
C_SCHOOL_LEVEL = 42  # 院校标签 (e.g. 985/211/双一流)
C_SCHOOL_LEVEL2 = 42  # 院校水平 (e.g. 原中央直属)


def map_batch(raw: str) -> str:
    """Normalize batch names to our internal labels."""
    mapping = {
        "本科批": "本科普通批",
        "本科提前批": "本科提前批",
        "专科批": "专科批",
        "专科提前批": "专科提前批",
    }
    # Fallback: remove special chars
    for k, v in mapping.items():
        if k in raw:
            return v
    return raw


def map_school_level(tag: str | None, level: str | None) -> str:
    """Extract a concise school level from the tags."""
    if not tag and not level:
        return "普通本科"
    combined = (tag or "") + " " + (level or "")
    if "985" in combined:
        return "985"
    if "211" in combined:
        return "211"
    if "双一流" in combined:
        return "双一流"
    if "普通" in combined:
        return "普通本科"
    if "职业" in combined or "专科" in combined or "职业技术" in combined:
        return "职业本科"
    return "普通本科"


def safe_int(val) -> int | None:
    """Convert to int, returning None for empty/invalid values."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def import_excel():
    """Main import routine."""
    src = EXCEL_PATH
    if not src.exists():
        print(f"[ERROR] Excel file not found: {src}")
        return False

    print(f"[INFO] Reading {src} ({src.stat().st_size / 1024 / 1024:.1f} MB)")
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["总表"]
    total_rows = ws.max_row - 3  # minus header rows
    print(f"[INFO] {total_rows} data rows in '总表'")

    # ── Backup existing DB ──────────────────────────────────────────
    if DB_PATH.exists():
        import shutil
        shutil.copy2(DB_PATH, BACKUP_PATH)
        print(f"[INFO] Backed up existing DB → {BACKUP_PATH}")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=OFF")  # faster bulk insert

    # ── Counters ───────────────────────────────────────────────────
    stats = {
        "plan_2025": 0,
        "plan_skipped": 0,
        "hist_2025_group": 0,
        "hist_2024_group": 0,
        "hist_2024_dup_skipped": 0,
    }

    # For 2024 admission_history, we need to aggregate by group
    # (school_code + group_code) → min min_score, min min_rank, sum plan
    agg_2024: dict[tuple, dict] = {}

    # ── Phase 1: Scan all rows, build plan + hist ─────────────────
    print("[INFO] Scanning rows...")

    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        if idx % 5000 == 0 and idx > 0:
            print(f"  ... processed {idx} rows")

        batch = str(row[C_BATCH]) if row[C_BATCH] else ""
        province = str(row[C_PROVINCE]) if row[C_PROVINCE] else "广东"
        subject = str(row[C_SUBJECT]) if row[C_SUBJECT] else ""
        school_code = str(row[C_SCHOOL_CODE]).strip() if row[C_SCHOOL_CODE] else ""
        school_name = str(row[C_SCHOOL_NAME]).strip() if row[C_SCHOOL_NAME] else ""
        group_code = str(row[C_GROUP_CODE]).strip() if row[C_GROUP_CODE] else ""

        if not school_code or not group_code or not batch:
            stats["plan_skipped"] += 1
            continue

        batch_norm = map_batch(batch)

        # ── School info ────────────────────────────────────────────
        city = str(row[C_CITY]).strip() if row[C_CITY] else ""
        level_tag = str(row[C_SCHOOL_LEVEL]).strip() if row[C_SCHOOL_LEVEL] else ""
        school_level = map_school_level(level_tag, "")

        # ═══════════════════════════════════════════════════════════
        # 1. enrollment_plan (2025)
        # ═══════════════════════════════════════════════════════════
        major_name = str(row[C_MAJOR_NAME]).strip() if row[C_MAJOR_NAME] else ""
        major_code = str(row[C_MAJOR_CODE]).strip() if row[C_MAJOR_CODE] else ""
        subject_req = str(row[C_SUBJECT_REQ]).strip() if row[C_SUBJECT_REQ] else ""
        plan_count = safe_int(row[C_PLAN_COUNT]) or 0
        group_name = str(row[C_GROUP_MAJORS]).strip() if row[C_GROUP_MAJORS] else ""

        if plan_count > 0 and major_name:
            conn.execute(
                """
                INSERT OR IGNORE INTO enrollment_plan
                    (year, province, batch, school_code, school_name,
                     group_code, group_name, major_code, major_name,
                     subject_requirement, subject_track, plan_count, school_level, city)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    2025, province, batch_norm,
                    school_code, school_name,
                    group_code, group_name,
                    major_code, major_name,
                    subject_req, subject,
                    plan_count,
                    school_level, city,
                ),
            )
            stats["plan_2025"] += 1

        # ═══════════════════════════════════════════════════════════
        # 2. admission_history (2025 group-level, cols 24-26)
        # ═══════════════════════════════════════════════════════════
        g25_min_score = safe_int(row[C_G25_MIN_SCORE])
        g25_min_rank = safe_int(row[C_G25_MIN_RANK])
        g25_admit = safe_int(row[C_G25_ADMIT])

        if g25_min_rank is not None and g25_min_rank > 0:
            conn.execute(
                """
                INSERT OR IGNORE INTO admission_history
                    (year, province, batch, school_code, school_name,
                     group_code, group_name, subject_track,
                     min_score, min_rank, plan_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    2025, province, batch_norm,
                    school_code, school_name,
                    group_code, group_name or major_name,
                    subject,
                    g25_min_score, g25_min_rank,
                    g25_admit or plan_count,
                ),
            )
            stats["hist_2025_group"] += 1

        # ═══════════════════════════════════════════════════════════
        # 3. admission_history (2024, aggregated from major-level)
        # ═══════════════════════════════════════════════════════════
        m24_min_score = safe_int(row[C_M24_MIN_SCORE])
        m24_min_rank = safe_int(row[C_M24_MIN_RANK])
        m24_admit = safe_int(row[C_M24_ADMIT])

        if m24_min_rank is not None and m24_min_rank > 0:
            key = (province, batch_norm, school_code, group_code)
            if key not in agg_2024:
                agg_2024[key] = {
                    "school_name": school_name,
                    "group_name": group_name or major_name,
                    "city": city,
                    "school_level": school_level,
                    "subject_track": subject,
                    "min_score": m24_min_score,
                    "min_rank": m24_min_rank,
                    "total_admit": m24_admit or 0,
                }
            else:
                existing = agg_2024[key]
                # Take the minimum (best) score/rank across majors in the group
                if m24_min_score is not None:
                    if existing["min_score"] is None or m24_min_score < existing["min_score"]:
                        existing["min_score"] = m24_min_score
                if m24_min_rank is not None:
                    if existing["min_rank"] is None or m24_min_rank < existing["min_rank"]:
                        existing["min_rank"] = m24_min_rank
                if m24_admit is not None:
                    existing["total_admit"] += m24_admit

    # ── Phase 2: Insert aggregated 2024 data ──────────────────────
    print(f"[INFO] Inserting {len(agg_2024)} aggregated 2024 groups...")
    for key, data in agg_2024.items():
        province, batch_norm, school_code, group_code = key
        conn.execute(
            """
            INSERT OR IGNORE INTO admission_history
                (year, province, batch, school_code, school_name,
                 group_code, group_name, subject_track,
                 min_score, min_rank, plan_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                2024, province, batch_norm,
                school_code, data["school_name"],
                group_code, data["group_name"],
                data.get("subject_track", ""),
                data["min_score"], data["min_rank"],
                data["total_admit"],
            ),
        )
        stats["hist_2024_group"] += 1

    # ── Commit & report ───────────────────────────────────────────
    conn.commit()
    conn.close()
    wb.close()

    print()
    print("=" * 60)
    print("Import Complete — Summary")
    print("=" * 60)
    print(f"  Enrollment Plan 2025:  {stats['plan_2025']} majors inserted")
    print(f"  Admission Hist 2025:   {stats['hist_2025_group']} groups inserted")
    print(f"  Admission Hist 2024:   {stats['hist_2024_group']} groups (aggregated)")
    print(f"  Skipped/invalid rows:  {stats['plan_skipped']}")
    print(f"  DB: {DB_PATH}")
    print()

    # Quick sanity check
    conn2 = sqlite3.connect(DB_PATH)
    for table in ("enrollment_plan", "admission_history"):
        count = conn2.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  [{table}] total rows: {count}")
        # Sample
        sample = conn2.execute(
            f"SELECT * FROM {table} LIMIT 3"
        ).fetchall()
        if sample:
            print(f"    Columns: {[d[0] for d in conn2.execute(f'PRAGMA table_info({table})').fetchall()]}")
            for s in sample:
                print(f"    → {s}")
    conn2.close()

    return True


if __name__ == "__main__":
    import_excel()

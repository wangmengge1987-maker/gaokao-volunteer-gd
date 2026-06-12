"""
Import 2026 Enrollment Plans from CSV/Excel into SQLite.

Usage:
  # From CSV
  python scripts/import_plan_2026.py --csv path/to/plan_2026.csv

  # From existing Excel (same format as 莫非凡)
  python scripts/import_plan_2026.py --excel "C:/path/to/database.xlsx"

  # Simulate demo data (for testing the impact analysis)
  python scripts/import_plan_2026.py --demo

CSV format (UTF-8):
  year,province,batch,school_code,school_name,group_code,group_name,major_code,major_name,subject_requirement,subject_track,plan_count,school_level,city

Demo mode generates plausible 2026 plan data by randomly adjusting 2025 plans
(±20%) for a subset of schools, allowing the impact analysis to be tested.
"""

import argparse
import csv
import random
import sqlite3
import sys
from pathlib import Path

DB_PATH = Path(__file__).resolve().parent.parent / "gaokao.db"


def backup_db():
    import shutil
    backup = DB_PATH.with_suffix(".db.2025_backup")
    if not backup.exists():
        shutil.copy2(DB_PATH, backup)
        print(f"[INFO] Backed up 2025 data → {backup}")
    else:
        print(f"[INFO] Backup already exists: {backup}")


def import_csv(csv_path: str):
    """Import 2026 plans from a CSV file."""
    path = Path(csv_path)
    if not path.exists():
        print(f"[ERROR] CSV not found: {path}")
        return False

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=OFF")

    count = 0
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        if not cols:
            print("[ERROR] Empty CSV")
            return False

        placeholders = ",".join("?" * len(cols))
        sql = f"INSERT OR IGNORE INTO enrollment_plan ({','.join(cols)}) VALUES ({placeholders})"

        for row in reader:
            values = tuple(row[c] for c in cols)
            conn.execute(sql, values)
            count += 1

    conn.commit()
    conn.close()

    print(f"[OK] Imported {count} rows of 2026 plan data from CSV")
    return True


def import_from_excel(excel_path: str):
    """Import 2026 plans from the same Excel format as the main workbook.
    
    Expects a sheet '2026计划' or uses '总表' and filters by year=2026.
    """
    import openpyxl

    path = Path(excel_path)
    if not path.exists():
        print(f"[ERROR] Excel not found: {path}")
        return False

    wb = openpyxl.load_workbook(str(path), data_only=True)

    # Try to find 2026 data
    sheet_name = None
    if "2026计划" in wb.sheetnames:
        sheet_name = "2026计划"
    elif "总表" in wb.sheetnames:
        sheet_name = "总表"
    else:
        print(f"[ERROR] No recognized sheet found. Available: {wb.sheetnames}")
        return False

    ws = wb[sheet_name]
    print(f"[INFO] Reading sheet '{sheet_name}' ({ws.max_row} rows)")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=OFF")

    # Column indices (matching import_data.py convention)
    C_YEAR = 0
    C_PROVINCE = 1
    C_BATCH = 2
    C_SUBJECT = 4
    C_SCHOOL_CODE = 5
    C_SCHOOL_NAME = 6
    C_GROUP_CODE = 7
    C_MAJOR_CODE = 9
    C_MAJOR_NAME = 11
    C_SUBJECT_REQ = 15
    C_PLAN_COUNT = 16
    C_GROUP_NAME = 19
    C_CITY = 39
    C_LEVEL_TAG = 42

    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        year = row[C_YEAR]
        if year is None or int(year) != 2026:
            continue

        batch = str(row[C_BATCH]) if row[C_BATCH] else "本科普通批"
        province = str(row[C_PROVINCE]) if row[C_PROVINCE] else "广东"
        subject = str(row[C_SUBJECT]) if row[C_SUBJECT] else ""
        school_code = str(row[C_SCHOOL_CODE]).strip() if row[C_SCHOOL_CODE] else ""
        school_name = str(row[C_SCHOOL_NAME]).strip() if row[C_SCHOOL_NAME] else ""
        group_code = str(row[C_GROUP_CODE]).strip() if row[C_GROUP_CODE] else ""
        major_name = str(row[C_MAJOR_NAME]).strip() if row[C_MAJOR_NAME] else ""
        major_code = str(row[C_MAJOR_CODE]).strip() if row[C_MAJOR_CODE] else ""
        subject_req = str(row[C_SUBJECT_REQ]).strip() if row[C_SUBJECT_REQ] else ""
        plan_count = int(row[C_PLAN_COUNT]) if row[C_PLAN_COUNT] else 0
        group_name = str(row[C_GROUP_NAME]).strip() if row[C_GROUP_NAME] else ""
        city = str(row[C_CITY]).strip() if row[C_CITY] else ""
        level_tag = str(row[C_LEVEL_TAG]).strip() if row[C_LEVEL_TAG] else ""

        # Map school level
        if "985" in level_tag:
            school_level = "985"
        elif "211" in level_tag:
            school_level = "211"
        elif "双一流" in level_tag:
            school_level = "双一流"
        else:
            school_level = "普通本科"

        if not school_code or not group_code:
            continue

        try:
            conn.execute(
                """
                INSERT OR IGNORE INTO enrollment_plan
                    (year, province, batch, school_code, school_name,
                     group_code, group_name, major_code, major_name,
                     subject_requirement, subject_track, plan_count, school_level, city)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    2026, province, batch,
                    school_code, school_name,
                    group_code, group_name,
                    major_code, major_name,
                    subject_req, subject,
                    plan_count,
                    school_level, city,
                ),
            )
            count += 1
        except Exception as e:
            print(f"  [WARN] Skip row: {e}")

    conn.commit()
    conn.close()
    wb.close()

    print(f"[OK] Imported {count} rows of 2026 plan data from Excel")
    return True


def generate_demo_data():
    """Generate demo 2026 plan data by randomly adjusting 2025 plans.
    
    Simulates realistic scenarios:
    - 中山大学 expands by 300 (famous example from user's description)
    - Some other top schools adjust by smaller amounts
    - Random ±20% for a subset of schools
    """
    print("[INFO] Generating demo 2026 plan data...")
    backup_db()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")

    # Get all 2025 plans
    rows = conn.execute(
        """
        SELECT school_code, school_name, group_code, group_name,
               subject_track, subject_requirement, school_level, city,
               SUM(plan_count) as total_plan
        FROM enrollment_plan
        WHERE year = 2025 AND batch = '本科普通批'
        GROUP BY school_code, group_code, subject_track
        """
    ).fetchall()
    plans_2025 = [dict(r) for r in rows]

    # Get admission ranks to identify top schools
    school_ranks = {}
    for p in plans_2025:
        key = (p["school_code"], p["group_code"], p["subject_track"])
        row = conn.execute(
            """
            SELECT min_rank FROM admission_history
            WHERE school_code = ? AND group_code = ? AND subject_track = ?
              AND min_rank IS NOT NULL
            ORDER BY year DESC LIMIT 1
            """,
            (p["school_code"], p["group_code"], p["subject_track"]),
        ).fetchone()
        if row:
            school_ranks[key] = row["min_rank"]

    # Define demo adjustments
    adjustments = {}
    # 中山大学 物理类 扩招 300（知名案例）
    adjustments[("10558", "10558215", "物理")] = +80  # 信息技术组
    adjustments[("10558", "10558216", "物理")] = +60  # 计算机组
    adjustments[("10558", "10558218", "物理")] = +50  # 临床医学
    adjustments[("10558", "10558217", "物理")] = +40  # 理工科组
    adjustments[("10558", "10558212", "物理")] = +30  # 人文社科
    adjustments[("10558", "10558213", "物理")] = +20
    adjustments[("10558", "10558214", "物理")] = +20
    adjustments[("10558", "10558219", "物理")] = +20  # 生化组

    # 华南理工 物理类 扩招 100
    adjustments[("10561", "201", "物理")] = +50
    adjustments[("10561", "202", "物理")] = +30
    adjustments[("10561", "203", "物理")] = +20

    # 广东工业大学 缩招 80（部分专业转移）
    adjustments[("11845", "401", "物理")] = -30
    adjustments[("11845", "402", "物理")] = -25
    adjustments[("11845", "403", "物理")] = -25

    # 深圳大学 扩招 60（新建学院）
    adjustments[("11105", "501", "物理")] = +30
    adjustments[("11105", "502", "物理")] = +30

    # 华南师范大学 扩招 40
    adjustments[("10574", "301", "物理")] = +40

    # Delete any existing 2026 demo data
    conn.execute("DELETE FROM enrollment_plan WHERE year = 2026")

    count_inserted = 0
    count_adjusted = 0
    for plan in plans_2025:
        key = (plan["school_code"], plan["group_code"], plan["subject_track"])
        adj = adjustments.get(key, 0)

        # For schools not in explicit adjustments, randomly adjust some
        if adj == 0 and random.random() < 0.10:  # 10% of other schools adjust
            adj = random.randint(-15, 25)

        new_plan = max(1, plan["total_plan"] + adj)
        if adj != 0:
            count_adjusted += 1

        # Insert ALL groups (with or without changes) so plan_changes
        # comparison works correctly for both years
        conn.execute(
            """
            INSERT INTO enrollment_plan
                (year, province, batch, school_code, school_name,
                 group_code, group_name, major_code, major_name,
                 subject_requirement, subject_track, plan_count, school_level, city)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                2026, "广东", "本科普通批",
                plan["school_code"], plan["school_name"],
                plan["group_code"], plan["group_name"] or "",
                "999999", f"2026计划({plan['group_code']})",
                plan["subject_requirement"] or "", plan["subject_track"],
                new_plan,
                plan["school_level"] or "普通本科", plan["city"] or "",
            ),
        )
        count_inserted += 1

    conn.commit()
    conn.close()

    print(f"[OK] Generated demo 2026 plan data across all schools")
    print(f"     Total groups inserted: {count_inserted}")
    print(f"     Groups with plan changes: {count_adjusted}")
    print()
    print("=" * 60)
    print("  Demo Scenario (模拟场景)")
    print("=" * 60)
    print("  中山大学 物理类 扩招 ~300 人（各专业组合计）")
    print("  华南理工大学 物理类 扩招 ~100 人")
    print("  广东工业大学 物理类 缩招 ~80 人")
    print("  深圳大学 物理类 扩招 ~60 人")
    print("  华南师范大学 物理类 扩招 ~40 人")
    print("  其他学校随机微调（±20~30 人）")
    print()
    print("  预期效果：")
    print("  中山大学录取位次从 ~8500 → ~8800（下移约300位）")
    print("  华工等下游学校位次同步下移 ~200~300 位")
    print("  = 原来'稳'的学校可能变成'冲'，'冲'的变得更有希望")
    print("=" * 60)
    return True


def show_status():
    """Print current data status."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.execute("SELECT COUNT(*) as cnt FROM enrollment_plan WHERE year=2026")
    count_2026 = cur.fetchone()["cnt"]
    cur = conn.execute("SELECT COUNT(*) as cnt FROM enrollment_plan WHERE year=2025")
    count_2025 = cur.fetchone()["cnt"]

    print(f"  2025 plans: {count_2025} rows")
    print(f"  2026 plans: {count_2026} rows")
    if count_2026 > 0:
        print("  [OK] 2026 plan data available")
    else:
        print("  [..] No 2026 plan data yet")

    if count_2026 > 0:
        print()
        rows = conn.execute("""
            SELECT school_name, SUM(plan_count) as total
            FROM enrollment_plan WHERE year=2026
            GROUP BY school_code ORDER BY total DESC LIMIT 10
        """).fetchall()
        print("  Top 10 by 2026 plan count:")
        for r in rows:
            print(f"    {r['school_name']}: {r['total']}")
    conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import 2026 enrollment plans")
    parser.add_argument("--csv", help="Path to CSV file")
    parser.add_argument("--excel", help="Path to Excel file")
    parser.add_argument("--demo", action="store_true", help="Generate demo data")
    parser.add_argument("--status", action="store_true", help="Show current status")

    args = parser.parse_args()

    if args.status:
        show_status()
        sys.exit(0)

    if args.csv:
        success = import_csv(args.csv)
    elif args.excel:
        success = import_from_excel(args.excel)
    elif args.demo:
        success = generate_demo_data()
    else:
        parser.print_help()
        print()
        print("No import mode specified. Use one of: --csv, --excel, --demo, --status")
        sys.exit(1)

    if success:
        print()
        show_status()
